import streamlit as st
import openpyxl
from io import BytesIO

st.title("Merge Name and Roll Columns in Excel")

# Upload Excel file
uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])
if uploaded_file:
    # Load workbook
    wb = openpyxl.load_workbook(uploaded_file)
    sheet = wb.active  # default to first sheet

    # Find first empty column
    max_col = sheet.max_column
    new_col = max_col + 1
    sheet.cell(row=1, column=new_col).value = "Name_Roll"

    # Merge Name (A) and Roll (B)
    for row in range(2, sheet.max_row + 1):
        name = sheet[f"A{row}"].value
        roll = sheet[f"B{row}"].value
        if name is not None and roll is not None:
            sheet.cell(row=row, column=new_col).value = f"{name} {roll}"

    # Save workbook to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Download button
    st.download_button(
        label="Download Merged Excel",
        data=output,
        file_name="merged_file.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
